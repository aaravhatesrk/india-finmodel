"""Builds a formatted, multi-sheet .xlsx financial model from scraped
company data: Overview, Income, Balance, Cash Flow, DCF Inputs."""
import io

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TEAL = "0F766E"
CORAL = "FF6B6B"
SAND = "F7EFE5"

HEADER_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color=TEAL, bold=True, size=14)
CURRENCY_FORMAT = '₹#,##0.00;[RED]-₹#,##0.00'


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")


def _autofit(ws, max_width=28):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_width, max(10, length + 2))


def _write_table(ws, records, start_row=3, currency_cols=None):
    """Write a list-of-dicts table starting at start_row, return last row used."""
    currency_cols = currency_cols or set()
    if not records:
        ws.cell(row=start_row, column=1, value="No data available")
        return start_row

    headers = list(records[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=header)
    _style_header_row(ws, row=start_row)

    for r_idx, record in enumerate(records, start=start_row + 1):
        for c_idx, header in enumerate(headers, start=1):
            value = record.get(header)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if header in currency_cols and isinstance(value, (int, float)):
                cell.number_format = CURRENCY_FORMAT

    last_row = start_row + len(records)
    if len(headers) >= 2:
        rule = ColorScaleRule(
            start_type="min", start_color="FFF1E6",
            end_type="max", end_color=CORAL,
        )
        col_letter = get_column_letter(len(headers))
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{last_row}", rule
        )
    return last_row


def _add_overview_sheet(wb, company):
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = company.get("name", company.get("symbol", "Company"))
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Symbol: {company.get('symbol', '')}"
    ws["A2"].font = Font(italic=True, color="555555")
    ws["A3"] = f"Source: {company.get('source_url', 'n/a')}"
    ws["A3"].font = Font(size=9, color="888888")

    ratios = company.get("ratios", {}) or {}
    row = 5
    ws.cell(row=row, column=1, value="Key Ratio")
    ws.cell(row=row, column=2, value="Value")
    _style_header_row(ws, row=row)
    for name, value in ratios.items():
        row += 1
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=value)

    _autofit(ws)
    ws.sheet_properties.tabColor = TEAL


def _add_financial_sheet(wb, title, records):
    ws = wb.create_sheet(title)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    last_row = _write_table(ws, records, start_row=3)

    if records and len(records[0].keys()) >= 2:
        n_cols = len(records[0].keys())
        chart = LineChart()
        chart.title = f"{title} trend"
        chart.style = 2
        chart.y_axis.title = "Value"
        chart.x_axis.title = records and list(records[0].keys())[0] or ""
        data = Reference(ws, min_col=2, max_col=n_cols, min_row=3, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=4, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, f"A{last_row + 3}")

    _autofit(ws)
    ws.sheet_properties.tabColor = CORAL
    return ws


def _add_dcf_sheet(wb, company, options):
    ws = wb.create_sheet("DCF Inputs")
    ws["A1"] = "DCF Inputs & Assumptions"
    ws["A1"].font = TITLE_FONT

    growth_rate = options.get("growth_rate", 0.10)
    discount_rate = options.get("discount_rate", 0.12)
    terminal_growth = options.get("terminal_growth", 0.04)
    projection_years = int(options.get("projection_years", 5))

    rows = [
        ("Revenue growth rate (Yr 1-5)", growth_rate),
        ("Discount rate (WACC)", discount_rate),
        ("Terminal growth rate", terminal_growth),
        ("Projection years", projection_years),
    ]
    row = 3
    ws.cell(row=row, column=1, value="Assumption")
    ws.cell(row=row, column=2, value="Value")
    _style_header_row(ws, row=row)
    for label, value in rows:
        row += 1
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if isinstance(value, float) and value < 1:
            cell.number_format = "0.0%"

    note_row = row + 2
    ws.cell(
        row=note_row,
        column=1,
        value=(
            "Note: populate base-year free cash flow from the Cash Flow sheet "
            "and extend this sheet with year-by-year projections as needed."
        ),
    )
    ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="888888")
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)

    _autofit(ws)
    ws.sheet_properties.tabColor = SAND


def build_excel_report(company: dict, options: dict | None = None) -> bytes:
    """Returns the raw .xlsx bytes for the given company data dict
    (as produced by scraper.get_company_data)."""
    options = options or {}
    wb = Workbook()

    _add_overview_sheet(wb, company)
    _add_financial_sheet(wb, "Income", company.get("profit_loss") or company.get("quarterly") or [])
    _add_financial_sheet(wb, "Balance", company.get("balance_sheet") or [])
    _add_financial_sheet(wb, "Cash Flow", company.get("cash_flow") or [])
    _add_dcf_sheet(wb, company, options)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
