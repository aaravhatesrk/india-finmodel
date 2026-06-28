import zipfile
from io import BytesIO

from openpyxl import load_workbook

from backend import excel_gen, powerbi_gen

SAMPLE_COMPANY = {
    "symbol": "TCS",
    "name": "Tata Consultancy Services",
    "source_url": "https://www.screener.in/company/TCS/consolidated/",
    "ratios": {"P/E": "28.5", "ROE": "45.2%"},
    "profit_loss": [
        {"Metric": "Revenue", "Mar 2023": 1500, "Mar 2024": 1700},
        {"Metric": "Net Profit", "Mar 2023": 300, "Mar 2024": 340},
    ],
    "balance_sheet": [{"Metric": "Total Assets", "Mar 2023": 5000, "Mar 2024": 5400}],
    "cash_flow": [{"Metric": "Operating Cash Flow", "Mar 2023": 400, "Mar 2024": 450}],
    "quarterly": [],
}


def test_build_excel_report_has_all_sheets():
    xlsx_bytes = excel_gen.build_excel_report(SAMPLE_COMPANY, {})
    wb = load_workbook(BytesIO(xlsx_bytes))
    assert wb.sheetnames == ["Overview", "Income", "Balance", "Cash Flow", "DCF Inputs"]


def test_build_excel_report_overview_has_name_and_ratios():
    xlsx_bytes = excel_gen.build_excel_report(SAMPLE_COMPANY, {})
    wb = load_workbook(BytesIO(xlsx_bytes))
    overview = wb["Overview"]
    assert overview["A1"].value == "Tata Consultancy Services"


def test_build_csv_contains_metrics():
    csv_bytes = powerbi_gen.build_csv(SAMPLE_COMPANY)
    text = csv_bytes.decode("utf-8")
    assert "Revenue" in text
    assert "Section,Metric,Period,Value" in text


def test_build_powerbi_package_zip_contents():
    zip_bytes = powerbi_gen.build_powerbi_package(SAMPLE_COMPANY, {})
    with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
        names = set(zf.namelist())
        assert names == {"data.csv", "placeholder.pbit", "INSTRUCTIONS.txt"}
